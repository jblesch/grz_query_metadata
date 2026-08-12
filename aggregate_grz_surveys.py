#!/usr/bin/env python3
"""
Aggregate per-GRZ survey reports into one workbook.

Takes the grz-survey-*.json files produced by survey_grz_metadata.py and writes
an .xlsx with one sheet per surveyed field: values down the rows, GRZs across
the columns, totals and share of all observations.

Optionally checks the free-text kit and instrument fields against a proposed
controlled vocabulary, so you can see what share of real values a candidate
enum would actually cover.

Usage
-----
    python aggregate_grz_surveys.py grz-survey-*.json -o schema-usage.xlsx

    # also measure coverage of the vocabularies proposed in
    # https://github.com/jblesch/MVGenomseq/pull/1 (GRZ/vocabularies/ on its dev branch)
    python aggregate_grz_surveys.py grz-survey-*.json -o schema-usage.xlsx \\
        --vocab labData.sequencerModel=GRZ/vocabularies/instrument-model.json \\
        --vocab labData.libraryPrepKit=GRZ/vocabularies/library-preparation-kit-retail-name.json

Requires: openpyxl
"""

from __future__ import annotations

import argparse
import collections
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEAD = Font(bold=True)
WARN = PatternFill("solid", fgColor="FFF2CC")
BAD = PatternFill("solid", fgColor="F8CECC")
UNUSED = PatternFill("solid", fgColor="DAE8FC")

# The metadata paths of the surveyed enum fields, so that a schema can be walked
# to the same places. Shipped alongside this script.
sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    from survey_grz_metadata import ENUM_FIELDS
except ImportError:  # only needed for --schema
    ENUM_FIELDS = {}


def normalise(v: str) -> str:
    """Loose key for comparing free text against a controlled value."""
    return re.sub(r"[^a-z0-9]+", "", v.lower())


# --------------------------------------------------------------------------
# Reading the declared enums out of the JSON Schema
# --------------------------------------------------------------------------

def _pointer(root: dict, frag: str):
    node = root
    for part in frag.strip("/").split("/"):
        if not part:
            continue
        part = part.replace("~1", "/").replace("~0", "~")
        node = node[part] if isinstance(node, dict) else node[int(part)]
    return node


def deref(node, root: dict, base: Path, depth: int = 0):
    """Follow $ref, local or to a sibling vocabulary file."""
    while isinstance(node, dict) and "$ref" in node and depth < 10:
        depth += 1
        ref = node["$ref"]
        target, _, frag = ref.partition("#")
        if not target:                      # local: "#/$defs/..."
            node = _pointer(root, frag)
        else:                               # file: "vocabularies/x.json"
            with open(base / target, encoding="utf-8") as fh:
                sub = json.load(fh)
            node = _pointer(sub, frag) if frag else sub
    return node


def declared_enum(root: dict, base: Path, path: str) -> list | None:
    """Walk a metadata path such as donors[]/labData[]/libraryType into the
    schema and return the enum declared there, or None if there isn't one."""
    node = root
    for seg in path.split("/"):
        is_array, key = seg.endswith("[]"), seg[:-2] if seg.endswith("[]") else seg
        node = deref(node, root, base)
        if not isinstance(node, dict):
            return None
        found = (node.get("properties") or {}).get(key)
        if found is None:                   # may sit inside allOf/anyOf/oneOf
            for comb in ("allOf", "anyOf", "oneOf"):
                for branch in node.get(comb) or []:
                    branch = deref(branch, root, base)
                    if isinstance(branch, dict) and key in (branch.get("properties") or {}):
                        found = branch["properties"][key]
                        break
                if found is not None:
                    break
        if found is None:
            return None
        node = found
        if is_array:
            node = deref(node, root, base)
            if not isinstance(node, dict) or "items" not in node:
                return None
            node = node["items"]
    node = deref(node, root, base)
    enum = node.get("enum") if isinstance(node, dict) else None
    return [str(v) for v in enum] if isinstance(enum, list) else None


def sheet_name(label: str) -> str:
    # Excel: max 31 chars, no []:*?/\
    s = re.sub(r"[\[\]:*?/\\]", "", label)
    return s[-31:]


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("reports", nargs="+", help="grz-survey-*.json files")
    ap.add_argument("-o", "--out", default="schema-usage.xlsx")
    ap.add_argument("--vocab", action="append", default=[],
                    help="FIELD=path/to/vocabulary.json, repeatable")
    ap.add_argument("--schema",
                    help="grz-schema.json to read the declared enums from. Adds a zero "
                         "row for every declared value that no GRZ ever used, and flags "
                         "used values that the schema does not declare.")
    args = ap.parse_args()

    reports = []
    for p in args.reports:
        with open(p, encoding="utf-8") as fh:
            r = json.load(fh)
        # grz_id is optional. Fall back to the filename, so that reports which
        # were produced without --grz-id still get one column each.
        gid = r.get("grz_id")
        r["_label"] = gid if gid and gid != "UNKNOWN" else Path(p).stem
        reports.append(r)
    reports.sort(key=lambda r: r["_label"])
    grzs = [r["_label"] for r in reports]
    if len(set(grzs)) != len(grzs):
        print(f"warning: duplicate report label among inputs: {grzs}", file=sys.stderr)

    vocabs = {}
    for spec in args.vocab:
        field, _, path = spec.partition("=")
        with open(path, encoding="utf-8") as fh:
            vocabs[field] = {normalise(v) for v in json.load(fh)["enum"]}

    declared: dict[str, list[str]] = {}
    if args.schema:
        if not ENUM_FIELDS:
            sys.exit("--schema requires survey_grz_metadata.py beside this script")
        spath = Path(args.schema)
        with open(spath, encoding="utf-8") as fh:
            schema_root = json.load(fh)
        for lbl, mpath in ENUM_FIELDS.items():
            found = declared_enum(schema_root, spath.parent, mpath)
            if found:
                declared[lbl] = found
        absent = sorted(set(ENUM_FIELDS) - set(declared))
        if absent:
            print(f"note: no enum declared in {spath.name} for {len(absent)} surveyed "
                  f"field(s): {', '.join(absent)}", file=sys.stderr)

    wb = Workbook()
    wb.remove(wb.active)

    # ---- summary -----------------------------------------------------
    ws = wb.create_sheet("summary")
    ws.append(["GRZ", "submissions in table", "with metadata", "unparseable",
               "script version", "generated", "free text redacted"])
    for c in ws[1]:
        c.font = HEAD
    for r in reports:
        ws.append([r["_label"], r["submissions_in_table"], r["submissions_with_metadata"],
                   r["submissions_unparseable"], r["script_version"], r["generated"],
                   "yes" if r.get("freetext_redacted") else "no"])
    ws.append([])
    ws.append(["TOTAL", sum(r["submissions_in_table"] for r in reports),
               sum(r["submissions_with_metadata"] for r in reports),
               sum(r["submissions_unparseable"] for r in reports)])
    ws[ws.max_row][0].font = HEAD
    versions = {r["script_version"] for r in reports}
    if len(versions) > 1:
        ws.append([])
        ws.append([f"WARNING: reports produced by different script versions: {sorted(versions)}"])
        ws[ws.max_row][0].fill = BAD

    # ---- one sheet per field ----------------------------------------
    sections = ("enum_fields", "freetext_fields", "derived")
    fields: list[tuple[str, str]] = []
    for sec in sections:
        for label in sorted({k for r in reports for k in r.get(sec, {})}):
            fields.append((sec, label))

    unused_by_field: dict[str, list[str]] = {}
    index = wb.create_sheet("index")
    index.append(["section", "field", "values used", "observations",
                  "declared in schema", "declared but NEVER used", "sheet"])
    for c in index[1]:
        c.font = HEAD

    for sec, label in fields:
        per_grz = {r["_label"]: r.get(sec, {}).get(label, {}).get("values", {}) for r in reports}
        observed = {v for d in per_grz.values() for v in d}
        decl = declared.get(label)
        # Declared values that nobody used still get a row, at zero. That absence
        # is the whole point of the survey, and it is invisible otherwise.
        values = sorted(observed | set(decl or []),
                        key=lambda v: (-sum(d.get(v, 0) for d in per_grz.values()), v))
        total_obs = sum(sum(d.values()) for d in per_grz.values())
        unused = [v for v in (decl or []) if v not in observed]
        name = sheet_name(label)
        index.append([sec, label, len(observed), total_obs,
                      len(decl) if decl else "", len(unused) if decl else "",
                      name])
        if decl and unused:
            index[index.max_row][5].fill = UNUSED

        ws = wb.create_sheet(name)
        header = ["value", *grzs, "TOTAL", "share"]
        if label in vocabs:
            header.append("in proposed vocabulary")
        if decl:
            header.append("declared in schema")
        ws.append(header)
        for c in ws[1]:
            c.font = HEAD
        ws.freeze_panes = "B2"

        for v in values:
            row = [v] + [per_grz[g].get(v, 0) for g in grzs]
            tot = sum(row[1:])
            row.append(tot)
            row.append(round(tot / total_obs, 4) if total_obs else 0)
            if label in vocabs:
                row.append("yes" if normalise(v) in vocabs[label] else "NO")
            if decl:
                # Not an error: submissions are validated when they arrive, so a
                # value missing from today's schema simply predates a change.
                row.append("yes" if v in decl else "no - predates this version")
            ws.append(row)
            cells = ws[ws.max_row]
            if decl and tot == 0:
                for c in cells:                       # declared but never used
                    c.fill = UNUSED
            elif label in vocabs and normalise(v) not in vocabs[label]:
                for c in cells:
                    c.fill = WARN

        ws.append([])
        ws.append(["TOTAL", *[sum(per_grz[g].values()) for g in grzs], total_obs])
        ws[ws.max_row][0].font = HEAD
        for i in range(1, len(header) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 46 if i == 1 else 16
        ws["A1"].alignment = Alignment(horizontal="left")

        if decl:
            unused_by_field[label] = unused
            ws.append([])
            ws.append([f"declared but never used ({len(unused)} of {len(decl)})",
                       ", ".join(unused) if unused else "none - every declared value occurs"])
            ws[ws.max_row][0].font = HEAD
            if unused:
                ws[ws.max_row][1].fill = UNUSED

        if label in vocabs:
            covered = sum(sum(per_grz[g].get(v, 0) for g in grzs)
                          for v in values if normalise(v) in vocabs[label])
            ws.append([])
            ws.append(["coverage by proposed vocabulary",
                       f"{covered}/{total_obs} observations",
                       f"{round(100 * covered / total_obs, 1) if total_obs else 0}%"])
            ws[ws.max_row][0].font = HEAD

    wb.save(args.out)
    print(f"wrote {args.out}: {len(fields)} field sheets from {len(reports)} GRZ report(s)")
    for field in vocabs:
        print(f"  vocabulary coverage checked for {field}")
    if declared:
        total_unused = sum(len(u) for u in unused_by_field.values())
        print(f"  checked {len(declared)} enum(s) against {Path(args.schema).name}: "
              f"{total_unused} declared value(s) never used")
        for lbl, u in sorted(unused_by_field.items()):
            if u:
                print(f"    {lbl}: {', '.join(u)}")


if __name__ == "__main__":
    main()
